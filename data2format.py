import os
import glob
import csv
import copy
from openpyxl import load_workbook

def main():
    target_dir = "./ble-pd-0C4314F46ABD_LOW"

    # ▼ マクロやVBAを使っていない通常のxlsxファイルなら keep_vba=True は不要です
    excel_file = os.path.join(target_dir, "RSSI可視化フォーマット.xlsx")
    wb = load_workbook(excel_file)
    
    # テンプレートシートを取得
    template_sheet = wb["format"]

    # CSVファイルの一覧を取得
    csv_files = sorted(glob.glob(os.path.join(target_dir, "*.csv")))
    total_files = len(csv_files)

    for i, csv_path in enumerate(csv_files, start=1):
        # CSVの名前から拡張子を除いたものを新規シート名に
        csv_name = os.path.splitext(os.path.basename(csv_path))[0]
        print(f"({i}/{total_files}) 処理中ファイル: {csv_name}")

        # 既に同名シートがあれば削除（再実行時の衝突回避）
        if csv_name in wb.sheetnames:
            del wb[csv_name]

        # テンプレートシートをコピー
        new_sheet = wb.copy_worksheet(template_sheet)
        new_sheet.title = csv_name
        
        # -------------------------------------------------------------
        # 条件付き書式を明示的に再コピー
        # -------------------------------------------------------------
        # copy_worksheet() は条件付き書式を完全にはコピーしないことがあるため、
        # ここで template_sheet の conditional_formatting を再度深いコピーで代入し直す
        new_sheet.conditional_formatting = copy.deepcopy(template_sheet.conditional_formatting)

        # -------------------------------------------------------------
        # CSVを読み込み → 数値(float)としてA3から書き込み
        # -------------------------------------------------------------
        with open(csv_path, mode="r", encoding="utf-8", newline="") as f:
            reader = csv.reader(f)
            for row_index, row_data in enumerate(reader, start=3):
                for col_index, value_str in enumerate(row_data, start=1):
                    # 全て数値データとのことなので float 変換
                    # 万一、空文字や数値以外が混じった場合は try/except で保護する方法もあります
                    float_value = float(value_str)
                    new_sheet.cell(row=row_index, column=col_index, value=float_value)

    # Excelを上書き保存
    wb.save(excel_file)
    print("全ファイルの処理が完了しました。")

if __name__ == "__main__":
    main()
